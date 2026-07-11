import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Financial Model V4"

# --- HELPER FUNCTIONS ---
def set_header(cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, color="FFFFFF")
    ws[cell].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

def set_label(cell, text, bold=False):
    ws[cell] = text
    if bold:
        ws[cell].font = Font(bold=True)

# --- 1. ПАРАМЕТРЫ СЕБЕСТОИМОСТИ ---
set_header("A1", "1. ПАРАМЕТРЫ СЕБЕСТОИМОСТИ (НАШИ ЗАТРАТЫ)")
ws.merge_cells("A1:C1")
set_label("A2", "Серверная база (Fix $/мес)")
ws["B2"] = 40
set_label("A3", "AI Скоринг (За 1 вакансию $)")
ws["B3"] = 1
set_label("A4", "Master Sales Nav (Цена $)")
ws["B4"] = 99
set_label("A5", "Master Sales Nav (Лимит вакансий)")
ws["B5"] = 20
set_label("A6", "Master LinkedHelper (Цена $)")
ws["B6"] = 15
set_label("A7", "Master LinkedHelper (Лимит вакансий)")
ws["B7"] = 20

# --- 2. МОДУЛИ И ПРАЙС-ЛИСТ ---
set_header("A9", "2. ПРАЙС-ЛИСТ НА МОДУЛИ (ЧТО МЫ ПРОДАЕМ КЛИЕНТУ ЗА 1 ВАКАНСИЮ)")
ws.merge_cells("A9:D9")
headers = ["Модуль", "Цена (Shared - Наш акк)", "Цена (BYOA - Свой акк)", "Логика"]
for i, h in enumerate(headers, 1):
    set_header(f"{get_column_letter(i)}10", h)

modules = [
    ("TG Sourcing", 19, 19, "ТГ бесплатный, разницы нет"),
    ("LinkedIn Sourcing", 69, 29, "Свой аккаунт = огромная скидка"),
    ("CRM & AI Scoring", 29, 29, "Используется наш API"),
    ("TG Outreach", 19, 19, "ТГ бесплатный"),
    ("LinkedIn Outreach", 49, 19, "Свой акк = дешевле и безопаснее")
]
for r, mod in enumerate(modules, 11):
    ws[f"A{r}"] = mod[0]
    ws[f"B{r}"] = mod[1]
    ws[f"C{r}"] = mod[2]
    ws[f"D{r}"] = mod[3]

# --- 3. КОНСТРУКТОР ТАРИФОВ И КЛИЕНТОВ ---
set_header("A18", "3. ПРОГНОЗ ПО КЛИЕНТАМ И ПАКЕТАМ (Укажите кол-во вакансий на пакет)")
ws.merge_cells("A18:L18")
headers = ["Название Пакета", "Кол-во Клиентов", "TG Src (вак)", "LI Src (Shared)", "LI Src (BYOA)", "AI Scoring", "TG Outreach", "LI Out (Shared)", "LI Out (BYOA)", "Цена 1 Пакета", "Итого Выручка с Пакета"]
for i, h in enumerate(headers, 1):
    set_header(f"{get_column_letter(i)}19", h)

packages = [
    ("Только TG Sourcing (Lite)", 5, 2, 0, 0, 2, 0, 0, 0),
    ("Full-Stack (Shared) (Pro)", 3, 4, 4, 0, 4, 4, 4, 0),
    ("Full-Stack (BYOA) (Ent)", 1, 10, 0, 10, 10, 10, 0, 10),
    ("Кастомный", 0, 0, 0, 0, 0, 0, 0, 0)
]

for r, pkg in enumerate(packages, 20):
    ws[f"A{r}"] = pkg[0]
    ws[f"B{r}"] = pkg[1] # Clients
    for c in range(2, 9):
        ws[f"{get_column_letter(c+1)}{r}"] = pkg[c]
    
    # Formula for Package Price: Sumproducts of chosen modules * prices
    # C*B11(TG) + D*B12(LISh) + E*C12(LIByoa) + F*B13(AI) + G*B14(TGO) + H*B15(LIOSh) + I*C15(LIOByoa)
    ws[f"J{r}"] = f"=C{r}*B11 + D{r}*B12 + E{r}*C12 + F{r}*B13 + G{r}*B14 + H{r}*B15 + I{r}*C15"
    # Total Rev
    ws[f"K{r}"] = f"=B{r}*J{r}"

set_label("A25", "ИТОГО:", bold=True)
ws["B25"] = "=SUM(B20:B23)" # Total clients
ws["K25"] = "=SUM(K20:K23)" # Total Revenue

# --- 4. РАСЧЕТ ИЗДЕРЖЕК И АГРЕГАЦИЯ ---
set_header("A27", "4. АГРЕГАЦИЯ И ИЗДЕРЖКИ (МАТЕМАТИКА)")
ws.merge_cells("A27:D27")

set_label("A28", "Всего LI Vacancies (Shared Sourcing)")
ws["B28"] = "=SUMPRODUCT(B20:B23, D20:D23)"

set_label("A29", "Всего LI Vacancies (Shared Outreach)")
ws["B29"] = "=SUMPRODUCT(B20:B23, H20:H23)"

set_label("A30", "Всего Вакансий для AI Скоринга")
ws["B30"] = "=SUMPRODUCT(B20:B23, F20:F23)"

set_label("A32", "Нужно аккаунтов Sales Nav (Shared)", bold=True)
ws["B32"] = "=ROUNDUP(B28/B5, 0)"

set_label("A33", "Нужно аккаунтов LinkedHelper (Shared)", bold=True)
ws["B33"] = "=ROUNDUP(B29/B7, 0)"

set_label("A35", "ОБЩИЕ РАСХОДЫ ($)", bold=True)
# Infra + (SN Accs * SN Price) + (LH Accs * LH Price) + (AI Vacs * AI Price)
ws["B35"] = "=B2 + (B32*B4) + (B33*B6) + (B30*B3)" 

# --- 5. ФИНАНСОВЫЙ ИТОГ ---
set_header("A38", "5. ФИНАНСОВЫЙ ИТОГ")
ws.merge_cells("A38:B38")

set_label("A39", "ОБЩАЯ ВЫРУЧКА", bold=True)
ws["B39"] = "=K25"
set_label("A40", "ОБЩИЕ РАСХОДЫ", bold=True)
ws["B40"] = "=B35"
set_label("A41", "ЧИСТАЯ ПРИБЫЛЬ", bold=True)
ws["B41"] = "=B39-B40"
set_label("A42", "Маржинальность", bold=True)
ws["B42"] = "=B41/B39"
ws["B42"].number_format = "0.0%"

set_label("A44", "Доля серверов (30%)")
ws["B44"] = "=B41*0.3"
set_label("A45", "Доля развитие (10%)")
ws["B45"] = "=B41*0.1"
set_label("A46", "Личная выручка (60%)")
ws["B46"] = "=B41*0.6"

# Width formatting
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 18
ws.column_dimensions["I"].width = 18
ws.column_dimensions["J"].width = 18
ws.column_dimensions["K"].width = 25

wb.save("/Users/imjustchilling/Desktop/sourcer/Financial_Model_Sourcer_V4_Modular.xlsx")
