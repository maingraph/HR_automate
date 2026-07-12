import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Financial Model V5"

# --- HELPER FUNCTIONS ---
def set_header(cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, color="FFFFFF")
    ws[cell].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

def set_label(cell, text, bold=False):
    ws[cell] = text
    if bold:
        ws[cell].font = Font(bold=True)

# --- 1. ПАРАМЕТРЫ СЕБЕСТОИМОСТИ ---
set_header("A1", "1. ПАРАМЕТРЫ СЕБЕСТОИМОСТИ И ЛИМИТЫ")
ws.merge_cells("A1:C1")
set_label("A2", "Серверная база (Fix $/мес)")
ws["B2"] = 40
set_label("A3", "AI Скоринг (За 1 вакансию $)")
ws["B3"] = 1
set_label("A4", "Рабочих дней в месяце (Для парсинга)")
ws["B4"] = 22
set_label("A5", "Master Sales Nav (Цена $)")
ws["B5"] = 99
set_label("A6", "Sales Nav (Скрапинг вакансий в ДЕНЬ на 1 акк)")
ws["B6"] = 1
set_label("A7", "Master LinkedHelper (Цена $)")
ws["B7"] = 15
set_label("A8", "LinkedHelper (Аутрич вакансий в ДЕНЬ на 1 акк)")
ws["B8"] = 1

# --- 2. МОДУЛИ И ПРАЙС-ЛИСТ ---
set_header("A10", "2. ПРАЙС-ЛИСТ НА МОДУЛИ (ЧТО МЫ ПРОДАЕМ КЛИЕНТУ ЗА 1 ВАКАНСИЮ)")
ws.merge_cells("A10:D10")
headers = ["Модуль", "Цена (Shared - Наш акк)", "Цена (BYOA - Свой акк)", "Логика"]
for i, h in enumerate(headers, 1):
    set_header(f"{get_column_letter(i)}11", h)

modules = [
    ("TG Sourcing", 19, 19, "ТГ бесплатный, разницы нет"),
    ("LinkedIn Sourcing", 69, 29, "Свой аккаунт = огромная скидка"),
    ("CRM & AI Scoring", 29, 29, "Используется наш API"),
    ("TG Outreach", 19, 19, "ТГ бесплатный"),
    ("LinkedIn Outreach", 49, 19, "Свой акк = дешевле и безопаснее")
]
for r, mod in enumerate(modules, 12):
    ws[f"A{r}"] = mod[0]
    ws[f"B{r}"] = mod[1]
    ws[f"C{r}"] = mod[2]
    ws[f"D{r}"] = mod[3]

# --- 3. КОНСТРУКТОР ТАРИФОВ И КЛИЕНТОВ ---
set_header("A19", "3. ПРОГНОЗ ПО КЛИЕНТАМ И ПАКЕТАМ (Укажите кол-во вакансий на пакет)")
ws.merge_cells("A19:L19")
headers = ["Название Пакета", "Кол-во Клиентов", "TG Src (вак)", "LI Src (Shared)", "LI Src (BYOA)", "AI Scoring", "TG Outreach", "LI Out (Shared)", "LI Out (BYOA)", "Цена 1 Пакета", "Итого Выручка с Пакета"]
for i, h in enumerate(headers, 1):
    set_header(f"{get_column_letter(i)}20", h)

packages = [
    ("Только TG Sourcing (Lite)", 5, 2, 0, 0, 2, 0, 0, 0),
    ("Full-Stack (Shared) (Pro)", 3, 4, 4, 0, 4, 4, 4, 0),
    ("Full-Stack (BYOA) (Ent)", 1, 10, 0, 10, 10, 10, 0, 10),
    ("Кастомный", 0, 0, 0, 0, 0, 0, 0, 0)
]

for r, pkg in enumerate(packages, 21):
    ws[f"A{r}"] = pkg[0]
    ws[f"B{r}"] = pkg[1] # Clients
    for c in range(2, 9):
        ws[f"{get_column_letter(c+1)}{r}"] = pkg[c]
    
    # Formula for Package Price
    ws[f"J{r}"] = f"=C{r}*B12 + D{r}*B13 + E{r}*C13 + F{r}*B14 + G{r}*B15 + H{r}*B16 + I{r}*C16"
    ws[f"K{r}"] = f"=B{r}*J{r}"

set_label("A26", "ИТОГО:", bold=True)
ws["B26"] = "=SUM(B21:B24)" # Total clients
ws["K26"] = "=SUM(K21:K24)" # Total Revenue

# --- 4. РАСЧЕТ ИЗДЕРЖЕК И АГРЕГАЦИЯ ---
set_header("A28", "4. АГРЕГАЦИЯ И ИЗДЕРЖКИ (ПОДСЧЕТ АККАУНТОВ)")
ws.merge_cells("A28:D28")

set_label("A29", "Всего LI Vacancies (Shared Sourcing)")
ws["B29"] = "=SUMPRODUCT(B21:B24, D21:D24)"

set_label("A30", "Всего LI Vacancies (Shared Outreach)")
ws["B30"] = "=SUMPRODUCT(B21:B24, H21:H24)"

set_label("A31", "Всего Вакансий для AI Скоринга")
ws["B31"] = "=SUMPRODUCT(B21:B24, F21:F24)"

set_label("A33", "Нужно аккаунтов Sales Nav (Shared)", bold=True)
# Formula: ROUNDUP( Total Vacancies / (Vacancies_Per_Day * Working_Days) )
ws["B33"] = "=ROUNDUP(B29/(B6*B4), 0)"

set_label("A34", "Нужно аккаунтов LinkedHelper (Shared)", bold=True)
ws["B34"] = "=ROUNDUP(B30/(B8*B4), 0)"

set_label("A36", "ОБЩИЕ РАСХОДЫ ($)", bold=True)
# Infra + (SN Accs * SN Price) + (LH Accs * LH Price) + (AI Vacs * AI Price)
ws["B36"] = "=B2 + (B33*B5) + (B34*B7) + (B31*B3)" 

# --- 5. ФИНАНСОВЫЙ ИТОГ ---
set_header("A39", "5. ФИНАНСОВЫЙ ИТОГ")
ws.merge_cells("A39:B39")

set_label("A40", "ОБЩАЯ ВЫРУЧКА", bold=True)
ws["B40"] = "=K26"
set_label("A41", "ОБЩИЕ РАСХОДЫ", bold=True)
ws["B41"] = "=B36"
set_label("A42", "ЧИСТАЯ ПРИБЫЛЬ", bold=True)
ws["B42"] = "=B40-B41"
set_label("A43", "Маржинальность", bold=True)
ws["B43"] = "=B42/B40"
ws["B43"].number_format = "0.0%"

set_label("A45", "Доля серверов (30%)")
ws["B45"] = "=B42*0.3"
set_label("A46", "Доля развитие (10%)")
ws["B46"] = "=B42*0.1"
set_label("A47", "Личная выручка (60%)")
ws["B47"] = "=B42*0.6"

# Width formatting
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 18
ws.column_dimensions["I"].width = 18
ws.column_dimensions["J"].width = 18
ws.column_dimensions["K"].width = 25

wb.save("/Users/imjustchilling/Desktop/sourcer/Financial_Model_Sourcer_V5_DailyLimits.xlsx")
