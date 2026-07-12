import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SaaS Economics V6"

def set_header(cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, color="FFFFFF")
    ws[cell].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

def set_label(cell, text, bold=False):
    ws[cell] = text
    if bold:
        ws[cell].font = Font(bold=True)

# --- 1. СРЕДНЕРЫНОЧНЫЕ ИЗДЕРЖКИ (RESEARCH BACKED) ---
set_header("A1", "1. СРЕДНЕРЫНОЧНЫЕ ИЗДЕРЖКИ (ИССЛЕДОВАНО НА РЫНКЕ SaaS)")
ws.merge_cells("A1:C1")

data_costs = [
    ("Базовая инфраструктура (Vercel, БД Supabase, Redis)", 40, "$/мес"),
    ("Серверная нагрузка (Compute/DB R&W) на 1 вакансию", 0.15, "$/вакансия"),
    ("Резидентные прокси (~25MB парсинга) на 1 вакансию", 0.10, "$/вакансия"),
    ("Токены AI (Gemini 1.5 - ~100k токенов) на 1 вакансию", 0.05, "$/вакансия"),
    ("Master Sales Navigator (Наш аккаунт)", 99, "$/мес"),
    ("Master LinkedHelper (Наш аккаунт)", 15, "$/мес"),
    ("Рабочих дней в месяце (Парсинг без выходных)", 22, "дней"),
    ("Sales Nav Лимит: парсинг вакансий в ДЕНЬ на 1 акк", 1, "вакансий/день"),
    ("LinkedHelper Лимит: аутрич вакансий в ДЕНЬ на 1 акк", 1, "вакансий/день")
]

for r, (label, val, unit) in enumerate(data_costs, 2):
    set_label(f"A{r}", label)
    ws[f"B{r}"] = val
    ws[f"C{r}"] = unit

set_label("A12", "ИТОГО: Переменные затраты на 1 вакансию (Без Master-аккаунтов)", bold=True)
ws["B12"] = "=B3+B4+B5"
ws["C12"] = "$/вакансия"

# --- 2. СТРАТЕГИЯ ЦЕНООБРАЗОВАНИЯ ---
set_header("A14", "2. МАРЖИНАЛЬНОСТЬ (ДЛЯ АВТО-РАСЧЕТА ЦЕНЫ КЛИЕНТУ)")
ws.merge_cells("A14:C14")
set_label("A15", "Целевая маржа для тарифов Shared (Мы платим за аккаунты)")
ws["B15"] = 0.85
ws["B15"].number_format = "0%"
set_label("A16", "Целевая маржа для тарифов BYOA (Клиент платит за свой Sales Nav)")
ws["B16"] = 0.95
ws["B16"].number_format = "0%"

# --- 3. ТАРИФНАЯ СЕТКА И КЛИЕНТЫ ---
set_header("A19", "3. ТАРИФНАЯ СЕТКА И АВТО-ЦЕНА (ВВЕДИТЕ ПРОГНОЗ ПО КЛИЕНТАМ)")
ws.merge_cells("A19:K19")
headers = ["Название Тарифа", "Лимит Вакансий (в месяц)", "Формат работы (Shared / BYOA)", "Доп. нагрузка (Outreach)", "СЕБЕСТОИМОСТЬ для нас ($)", "РЕКОМЕНДУЕМАЯ ЦЕНА КЛИЕНТУ ($)", "ПРИБЫЛЬ с 1 клиента ($)", "ПРОГНОЗ: Кол-во Клиентов", "ИТОГО Выручка ($)"]
for i, h in enumerate(headers, 1):
    set_header(f"{get_column_letter(i)}20", h)
    ws.column_dimensions[get_column_letter(i)].width = 22

ws.column_dimensions["A"].width = 30

tiers = [
    ("Lite (Только TG, Без Sales Nav)", 5, "Shared (ТГ бесплатный)", "Нет", "=B21*B$12", "=E21/(1-B$15)", "=F21-E21", 5, "=F21*H21"),
    ("Pro (Full-stack Shared аккаунты)", 10, "Shared", "Да (LinkedHelper)", "=(B22*B$12) + (B22/(B$8*B$9))*B$6 + (B22/(B$8*B$10))*B$7", "=E22/(1-B$15)", "=F22-E22", 3, "=F22*H22"),
    ("Enterprise (Full-stack BYOA)", 30, "BYOA (Свои аккаунты)", "Да (Свой LinkedHelper)", "=B23*B$12", "=E23/(1-B$16)", "=F23-E23", 1, "=F23*H23")
]

for r, t in enumerate(tiers, 21):
    ws[f"A{r}"] = t[0]
    ws[f"B{r}"] = t[1]
    ws[f"C{r}"] = t[2]
    ws[f"D{r}"] = t[3]
    ws[f"E{r}"] = t[4] # Cost
    ws[f"F{r}"] = t[5] # Price
    ws[f"F{r}"].number_format = "0"
    ws[f"G{r}"] = t[6] # Profit
    ws[f"H{r}"] = t[7] # Clients
    ws[f"I{r}"] = t[8] # Total Rev

# Total Clients
set_label("G25", "ВСЕГО КЛИЕНТОВ:", bold=True)
ws["H25"] = "=SUM(H21:H23)"

# --- 4. РЕАЛЬНЫЕ РАСХОДЫ И ИТОГИ БИЗНЕСА ---
set_header("A27", "4. ФИНАНСОВЫЕ ИТОГИ И МАСШТАБИРОВАНИЕ ИНФРАСТРУКТУРЫ")
ws.merge_cells("A27:D27")

set_label("A28", "Суммарно Вакансий в месяц (Все клиенты)")
ws["B28"] = "=SUMPRODUCT(B21:B23, H21:H23)"

set_label("A29", "Суммарно Shared Вакансий (Только Pro)")
ws["B29"] = "=B22*H22"

set_label("A30", "Требуется Master-аккаунтов Sales Nav", bold=True)
ws["B30"] = "=ROUNDUP(B29/(B8*B9), 0)"

set_label("A31", "Требуется Master-аккаунтов LinkedHelper", bold=True)
ws["B31"] = "=ROUNDUP(B29/(B8*B10), 0)"

set_label("A33", "ВЫРУЧКА БИЗНЕСА ($/мес)", bold=True)
ws["B33"] = "=SUM(I21:I23)"

set_label("A34", "ФАКТИЧЕСКИЕ РАСХОДЫ ($/мес)", bold=True)
# Base + (Total Vacs * VarCost) + (Accs * AccPrice)
ws["B34"] = "=B2 + (B28*B12) + (B30*B6) + (B31*B7)"

set_label("A35", "ЧИСТАЯ ПРИБЫЛЬ ($/мес)", bold=True)
ws["B35"] = "=B33-B34"

set_label("A36", "Реальная маржинальность бизнеса", bold=True)
ws["B36"] = "=B35/B33"
ws["B36"].number_format = "0.0%"

set_label("A38", "РАСПРЕДЕЛЕНИЕ ПРИБЫЛИ:")
ws.merge_cells("A38:B38")
set_label("A39", "На сервера и резерв (30%)")
ws["B39"] = "=B35*0.3"
set_label("A40", "На развитие (10%)")
ws["B40"] = "=B35*0.1"
set_label("A41", "Личная выручка (60%)")
ws["B41"] = "=B35*0.6"

ws.column_dimensions["A"].width = 50

wb.save("/Users/imjustchilling/Desktop/sourcer/Financial_Model_Sourcer_V6_Automated.xlsx")
